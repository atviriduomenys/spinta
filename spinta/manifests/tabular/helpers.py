from __future__ import annotations

import csv
import pathlib
import logging
import textwrap
import types
import uuid
from operator import itemgetter
from itertools import zip_longest
from typing import Any
from typing import Callable
from typing import Dict
from typing import IO
from typing import Iterable
from typing import Iterator
from typing import List
from typing import NamedTuple
from typing import Optional
from typing import Set
from typing import Tuple
from typing import TypeVar
from typing import Union
from typing import cast

import openpyxl
import xlsxwriter
from lark import ParseError
from tabulate import tabulate

from spinta import commands
from spinta import spyna
from spinta.backends import Backend
from spinta.backends.constants import BackendOrigin
from spinta.components import Context, Base, PrepareGiven
from spinta.datasets.components import Resource, Param
from spinta.dimensions.comments.components import Comment
from spinta.dimensions.enum.components import EnumItem
from spinta.components import Model
from spinta.components import Namespace
from spinta.components import Property
from spinta.core.enums import Access
from spinta.core.ufuncs import unparse
from spinta.datasets.components import Dataset
from spinta.dimensions.enum.components import Enums
from spinta.dimensions.lang.components import LangData
from spinta.dimensions.prefix.components import UriPrefix
from spinta.exceptions import MultipleErrors, InvalidBackRefReferenceAmount, DataTypeCannotBeUsedForNesting, \
    NestedDataTypeMismatch
from spinta.exceptions import PropertyNotFound
from spinta.manifests.components import Manifest
from spinta.manifests.helpers import load_manifest_nodes
from spinta.manifests.tabular.components import ACCESS, URI
from spinta.manifests.tabular.components import BackendRow
from spinta.manifests.tabular.components import BaseRow
from spinta.manifests.tabular.components import CommentData
from spinta.manifests.tabular.components import DESCRIPTION
from spinta.manifests.tabular.components import DatasetRow
from spinta.manifests.tabular.components import ParamRow
from spinta.manifests.tabular.components import EnumRow
from spinta.manifests.tabular.components import ID
from spinta.manifests.tabular.components import MANIFEST_COLUMNS
from spinta.manifests.tabular.components import ManifestColumn
from spinta.manifests.tabular.components import ManifestRow
from spinta.manifests.tabular.components import ManifestTableRow
from spinta.manifests.tabular.components import ModelRow
from spinta.manifests.tabular.components import PREPARE
from spinta.manifests.tabular.components import PROPERTY
from spinta.manifests.tabular.components import PrefixRow
from spinta.manifests.tabular.components import PropertyRow
from spinta.manifests.tabular.components import REF
from spinta.manifests.tabular.components import ResourceRow
from spinta.manifests.tabular.components import SOURCE
from spinta.manifests.tabular.components import TITLE
from spinta.manifests.tabular.components import LEVEL
from spinta.manifests.tabular.components import TabularFormat
from spinta.manifests.tabular.constants import DATASET
from spinta.manifests.tabular.constants import DataTypeEnum
from spinta.manifests.tabular.formats.gsheets import read_gsheets_manifest
from spinta.spyna import SpynaAST
from spinta.types.datatype import Ref, DataType, Denorm, Inherit, ExternalRef, BackRef, ArrayBackRef, Array, Object
from spinta.utils.data import take
from spinta.utils.schema import NA
from spinta.utils.schema import NotAvailable
from spinta.types.text.components import Text

log = logging.getLogger(__name__)

ParsedRow = Tuple[int, Dict[str, Any]]

MAIN_DIMENSIONS = [
    'dataset',
    'resource',
    'base',
    'model',
    'property',
]

EXTRA_DIMENSIONS = [
    '',
    'prefix',
    'enum',
    'param',
    'comment',
    'ns',
    'lang',
    'unique',
]

ALLOWED_PARTIAL_TYPES = [
    DataTypeEnum._OBJECT.value,
    DataTypeEnum._PARTIAL.value,
    DataTypeEnum.REF.value,
    DataTypeEnum.BACKREF.value,
]

ALLOWED_ARRAY_TYPES = [
    DataTypeEnum.ARRAY.value,
    DataTypeEnum._PARTIAL_ARRAY.value,
    DataTypeEnum._ARRAY_BACKREF.value,
]

ALLOWED_NESTING_TYPES = ALLOWED_PARTIAL_TYPES + ALLOWED_ARRAY_TYPES + [
    DataTypeEnum.TEXT.value
]


class TabularManifestError(Exception):
    pass


def _detect_header(
    path: Optional[str],
    line: int,  # Line number
    row: Iterable[str],
) -> List[str]:
    header = [h.strip().lower() for h in row]
    unknown_columns = set(header[:len(DATASET)]) - set(DATASET)
    if unknown_columns:
        unknown_columns = ', '.join(sorted(unknown_columns, key=header.index))
        raise TabularManifestError(
            f"{path}:{line}: Unknown columns: {unknown_columns}."
        )
    return header


def _detect_dimension(
    path: Optional[pathlib.Path],
    line: str,  # Line number with a prefix (depends on manifest format)
    row: Dict[str, str],
) -> Optional[str]:
    dimensions = [k for k in MAIN_DIMENSIONS if row[k]]

    if len(dimensions) == 1:
        return dimensions[0]

    if len(dimensions) > 1:
        dimensions = ', '.join([f'{k}: {row[k]}' for k in dimensions])
        raise TabularManifestError(
            f"{path}:{line}: In one row only single dimension can be used, "
            f"but found more than one: {dimensions}"
        )

    if row['type']:
        if row['type'] not in EXTRA_DIMENSIONS:
            raise TabularManifestError(
                f"{path}:{line}:type: Unknown additional dimension name "
                f"{row['type']}."
            )
        return row['type']

    return ''


def _parse_spyna(
    reader: TabularReader,
    formula: str,
) -> Union[SpynaAST, NotAvailable, None]:
    if formula:
        try:
            return spyna.parse(formula)
        except ParseError as e:
            reader.error(f"Error while parsing formula {formula!r}:\n{e}")
    return NA


class TabularReader:
    state: State
    path: str
    line: str
    type: str
    name: str
    data: ManifestRow               # Used when `appendable` is False
    rows: List[Dict[str, Any]]      # Used when `appendable` is True
    appendable: bool = False        # Tells if reader is appendable.
    allow_updates: bool = False     # Tells if manifest supports structure updates

    def __init__(
        self,
        state: State,
        path: str,
        line: str,
        allow_updates: bool = False
    ):
        self.state = state
        self.path = path
        self.line = line
        self.data = {}
        self.rows = []
        self.allow_updates = allow_updates

    def __str__(self):
        return f"<{type(self).__name__} name={self.name!r}>"

    def read(self, row: Dict[str, str]) -> None:
        raise NotImplementedError

    def append(self, row: Dict[str, str]) -> None:
        if not self.allow_updates:
            if any(row.values()):
                self.error(
                    f"Updates are not supported in context of {self.type!r}."
                )

    def release(self, reader: TabularReader = None) -> bool:
        raise NotImplementedError

    def items(self) -> Iterator[ParsedRow]:
        if self.appendable:
            for data in self.rows:
                yield self.line, data
        else:
            yield self.line, self.data

    def enter(self) -> None:
        raise NotImplementedError

    def leave(self) -> None:
        raise NotImplementedError

    def error(self, message: str) -> None:
        raise TabularManifestError(f"{self.path}:{self.line}: {message}")


class ManifestReader(TabularReader):
    type: str = 'manifest'
    datasets: Set[str]
    namespaces: Set[str]
    data: ManifestTableRow

    def read(self, row: ManifestRow) -> None:
        self.name = str(self.path)
        self.data = {
            'type': 'manifest',
        }

    def release(self, reader: TabularReader = None) -> bool:
        return reader is None

    def enter(self) -> None:
        self.datasets = set()
        self.namespaces = set()
        self.state.manifest = self

    def leave(self) -> None:
        self.state.manifest = None


class DatasetReader(TabularReader):
    type: str = 'dataset'
    data: DatasetRow

    def read(self, row: Dict[str, str]) -> None:
        self.name = row['dataset']

        if self.name == '/':
            self.data = {}
        else:
            if row['dataset'] in self.state.manifest.datasets:
                self.error("Dataset already defined.")

            self.data = {
                'type': 'dataset',
                'id': row['id'],
                'name': row['dataset'],
                'source': row['source'],
                'level': row['level'],
                'access': row['access'],
                'title': row['title'],
                'description': row['description'],
                'given_name': row['dataset'],
                'resources': {},
            }

    def release(self, reader: TabularReader = None) -> bool:
        return reader is None or isinstance(reader, (
            ManifestReader,
            DatasetReader,
        )) or (
            isinstance(reader, (ResourceReader, ModelReader)) and
            self.name == '/'
        )

    def enter(self) -> None:
        self.state.dataset = self

    def leave(self) -> None:
        self.state.dataset = None


class ResourceReader(TabularReader):
    type: str = 'resource'
    data: Union[BackendRow, ResourceRow]

    def read(self, row: Dict[str, str]) -> None:
        self.name = row['resource']

        if self.name == '/':
            self.data = {}
        else:
            if self.state.dataset is None:
                self.read_backend(row)
            else:
                self.read_resource(row)

    def read_backend(self, row: Dict[str, str]) -> None:
        # Backends will be loaded using
        # `spinta.manifests.helpers._load_manifest_backends`.

        if 'backends' not in self.state.manifest.data:
            self.state.manifest.data['backends'] = {}
        backends = self.state.manifest.data['backends']

        if self.name in backends:
            self.error(
                f"Backend {self.name!r} with the same name already defined."
            )

        self.data = {
            'id': row['id'],
            'type': row['type'],
            'name': self.name,
            'dsn': row['source'],
            'title': row['title'],
            'description': row['description'],
        }

        backends[self.name] = self.data

    def read_resource(self, row: Dict[str, str]) -> None:
        dataset = _get_state_obj(self.state.dataset)
        dataset = dataset.data if dataset else None

        if self.name in dataset['resources']:
            self.error("Resource with the same name already defined in ")

        self.data = {
            'id': row['id'],
            'type': row['type'],
            'backend': row['ref'],
            'external': row['source'],
            'prepare': _parse_spyna(self, row[PREPARE]),
            'level': row['level'],
            'access': row['access'],
            'title': row['title'],
            'description': row['description'],
            'given_name': self.name,
        }

        dataset['resources'][self.name] = self.data

    def release(self, reader: TabularReader = None) -> bool:
        if self.state.dataset is None:
            return True
        return reader is None or isinstance(reader, (
            ManifestReader,
            DatasetReader,
            ResourceReader,
            EnumReader,
            PrefixReader
        )) or (isinstance(reader, ModelReader) and self.name == '/')

    def enter(self) -> None:
        self.state.resource = self

    def leave(self) -> None:
        self.state.resource = None


class BaseReader(TabularReader):
    type: str = 'base'
    data: BaseRow

    def read(self, row: Dict[str, str]) -> None:
        self.name = row['base']

        if self.name == '/':
            self.data = {}
        else:
            dataset = _get_state_obj(self.state.dataset)
            dataset = dataset.data if dataset else None

            self.data = {
                'id': row['id'],
                'name': self.name,
                'model': get_relative_model_name(dataset, row['base']),
                'pk': (
                    [x.strip() for x in row['ref'].split(',')]
                    if row['ref'] else []
                ),
                'level': row['level']
            }

    def release(self, reader: TabularReader = None) -> bool:
        return reader is None or isinstance(reader, (
            ManifestReader,
            DatasetReader,
            ResourceReader,
            BaseReader,
        )) or (isinstance(reader, ModelReader) and self.name == '/')

    def enter(self) -> None:
        self.state.base = self

    def leave(self) -> None:
        self.state.base = None


class ModelReader(TabularReader):
    type: str = 'model'
    data: ModelRow

    def read(self, row: Dict[str, str]) -> None:
        dataset = _get_state_obj(self.state.dataset)
        resource = _get_state_obj(self.state.resource)
        base = _get_state_obj(self.state.base)
        name = get_relative_model_name(
            dataset.data if dataset else None,
            row['model'],
        )

        if self.state.rename_duplicates:
            dup = 1
            _name = name
            while _name in self.state.models:
                _name = f'{name}_{dup}'
                dup += 1
            name = _name
        elif name in self.state.models:
            self.error(f"Model {name!r} with the same name is already defined.")

        self.name = name

        self.data = {
            'type': 'model',
            'id': row['id'],
            'name': name,
            'base': {
                'id': base.data["id"],
                'name': base.name,
                'parent': base.data['model'],
                'pk': base.data['pk'],
                'level': base.data['level']
            } if base and base.data else None,
            'level': row['level'],
            'access': row['access'],
            'title': row['title'],
            'description': row['description'],
            'properties': {},
            'uri': row['uri'],
            'unique': [([x.strip() for x in row['ref'].split(',')])] if row['ref'] else [],
            'external': {
                'dataset': dataset.name if dataset else '',
                'resource': resource.name if dataset and resource else '',
                'pk': (
                    [x.strip() for x in row['ref'].split(',')]
                    if row['ref'] else []
                ),
                'name': row['source'],
                'prepare': _parse_spyna(self, row[PREPARE]),
            },
            'given_name': name,
        }
        if resource and not dataset:
            self.data['backend'] = resource.name

    def release(self, reader: TabularReader = None) -> bool:
        return reader is None or isinstance(reader, (
            ManifestReader,
            DatasetReader,
            ResourceReader,
            BaseReader,
            ModelReader,
        ))

    def enter(self) -> None:
        self.state.model = self
        self.state.models.add(self.name)

    def leave(self) -> None:
        self.state.model = None


def _parse_property_ref(ref: str) -> Tuple[str, List[str]]:
    if '[' in ref:
        ref = ref.rstrip(']')
        ref_model, ref_props = ref.split('[', 1)
        ref_props = [p.strip() for p in ref_props.split(',')]
    else:
        ref_model = ref
        ref_props = []
    return ref_model, ref_props


def _parse_dtype_string(dtype: str) -> dict:
    args = []
    error = None
    required = unique = False
    invalid_args = []

    if '(' in dtype:
        dtype, args = dtype.split('(', 1)
        args, additional_args = args.split(')', 1)
        args = args.strip().rstrip(')')
        args = [a.strip() for a in args.split(',')]
    else:
        if len(dtype.split(None, 1)) > 1:
            dtype, additional_args = dtype.split(None, 1)
        else:
            additional_args = ""

    if additional_args:
        for arg in additional_args.split(None):
            if arg == 'required':
                required = True
            elif arg == 'unique':
                unique = True
            else:
                invalid_args.append(arg)
        if invalid_args:
            error = f'Invalid type arguments: {", ".join(invalid_args)}.'

    return {
        'type': dtype,
        'type_args': args,
        'required': required,
        'unique': unique,
        'error': error,
    }


def _get_type_repr(dtype: List[DataType, str]):
    if isinstance(dtype, DataType):
        args = ''
        required = ' required' if dtype.required else ''
        unique = ' unique' if dtype.unique else ''

        model = dtype.prop.model
        if model.external and model.external.unknown_primary_key is False:
            if len(model.external.pkeys) == 1 and dtype.prop in model.external.pkeys:
                unique = ''
        if dtype.type_args:
            args = ', '.join(dtype.type_args)
            args = f'({args})'
        dtype_name = dtype.name if not isinstance(dtype, (
            Denorm, Inherit, ExternalRef, ArrayBackRef)) else dtype.get_type_repr()
        return f'{dtype_name}{args}{required}{unique}'
    else:
        args = ''
        required = ' required' if 'required' in dtype else ''
        unique = ' unique' if 'unique' in dtype else ''
        additional_args = []
        if '(' in dtype:
            dtype, args = dtype.split('(', 1)
            args, additional_args = args.split(')', 1)
            args = args.strip().rstrip(')')
            args = [a.strip() for a in args.split(',')]
            args = ', '.join(args)
            args = f'({args})'
        else:
            if len(dtype.split(None, 1)) > 1:
                dtype, additional_args = dtype.split(None, 1)
            else:
                dtype = dtype.strip(' ')
        if additional_args:
            if [additional_arg for additional_arg in additional_args.split(' ') if additional_arg not in [
                required.strip(' '), unique.strip(' ')]]:
                raise TabularManifestError
        return f'{dtype}{args}{required}{unique}'


def combine_source_prepare(source, prepare):
    result = prepare
    if source:
        if result:
            split = result.split('.')
            formula = split[0]
            split_formula = formula.split('(')
            formatted = source.replace('"', '\\"').replace("'", "\\'")

            if split_formula[1].strip() == ')':
                reconstructed = f'{split_formula[0]}("{formatted}"{"(".join(split_formula[1:])}'
            else:
                reconstructed = f'{split_formula[0]}("{formatted}", {"(".join(split_formula[1:])}'
            split[0] = reconstructed
            result = '.'.join(split)
        else:
            result = source
    return result


class PropertyReader(TabularReader):
    type: str = 'property'
    data: PropertyRow
    enums: Set[str]

    def read(self, row: Dict[str, str]) -> None:
        self.path_to_current_prop = self._path_to_current_prop(row['property'])
        complete_structure, parent_structure, prop_name = _extract_and_create_parent_data(self, row, row['property'])

        prop_data = _handle_datatype(self, row)
        prop_name = _combine_parent_with_prop(
            prop_name,
            prop_data,
            parent_structure,
            complete_structure
        )
        self.data = complete_structure
        self.name = prop_name
        self.state.model.data['properties'][prop_name] = self.data

    def append(self, row: Dict[str, str]) -> None:
        if not row['property']:
            result = combine_source_prepare(row['source'], row['prepare'])
            if not result:
                return
            self._append_prepare(row, result)

    def release(self, reader: TabularReader = None) -> bool:
        return reader is None or isinstance(reader, (
            ManifestReader,
            DatasetReader,
            ResourceReader,
            BaseReader,
            ModelReader,
            PropertyReader,
            UniqueReader
        ))

    def enter(self) -> None:
        self.state.prop = self

    def leave(self) -> None:
        self._parse_prepare()
        if 'external' in self.data:
            self.data['external']['prepare'] = self.data.pop('prepare') if 'prepare' in self.data else NA

        self.state.prop = None

    def _parse_prepare(self):
        if "prepare" in self.data:
            self.data["prepare"] = _parse_spyna(self, self.data["prepare"])

    def _append_prepare(self, row: Dict[str, str], prepare: str):
        if "prepare" in self.data:
            prep = self.data["prepare"]
            self.data["prepare"] = f'{prep}.{prepare}' if prep else prepare
        self.data['prepare_given'].append(
            PrepareGiven(
                appended=True,
                source=row['source'],
                prepare=row['prepare']
            )
        )

    def _path_to_current_prop(self, prop_given_name: str) -> str:
        STR_PROPERTIES = 'properties'
        parts = prop_given_name.split('.')[1:]
        result = '.'.join(
            part + ('.' + STR_PROPERTIES if i < len(parts) - 1 else '')
            for i, part in enumerate(parts)
        )
        return STR_PROPERTIES + '.' + result if result else ''


def _initial_normal_property_schema(given_name: str, dtype: dict, row: dict):
    return {
        'id': row.get('id'),
        'type': dtype['type'],
        'type_args': dtype['type_args'],
        'prepare': row.get(PREPARE),
        'level': row.get(LEVEL),
        'access': row.get(ACCESS),
        'uri': row.get(URI),
        'title': row.get(TITLE),
        'description': row.get(DESCRIPTION),
        'required': dtype['required'],
        'unique': dtype['unique'],
        'given_name': given_name,
        'prepare_given': [],
        'explicitly_given': True
    }


def _initial_array_property_schema(given_name: str, dtype: dict, row: dict):
    result = _initial_normal_property_schema(given_name, dtype, row)
    result['items'] = {}
    return result


def _initial_partial_property_schema(given_name: str, dtype: dict, row: dict):
    result = _initial_normal_property_schema(given_name, dtype, row)
    result['properties'] = {}
    return result


def _initial_text_property_schema(given_name: str, dtype: dict, row: dict):
    result = _initial_normal_property_schema(given_name, dtype, row)
    result['langs'] = {}
    return result


def _datatype_handler(reader: PropertyReader, row: dict, initial_data_loader: Callable[[str, dict, dict], dict]):
    given_name = row['property']
    reader.name = _clean_up_prop_name(row['property'].split('.')[-1])

    if reader.state.model is None:
        context = reader.state.stack[-1]
        reader.error(
            f"Property {reader.name!r} must be defined in a model context. "
            f"Now it is defined in {context.name!r} {context.type} context."
        )
    _check_if_property_already_set(reader, row, given_name)
    dtype = _get_type_repr(row['type'])
    dtype = _parse_dtype_string(dtype)
    if dtype['error']:
        reader.error(
            dtype['error']
        )

    if reader.state.base and not dtype['type']:
        dtype['type'] = 'inherit'
    elif '.' in given_name and not dtype['type']:
        dtype['type'] = 'denorm'

    new_data = initial_data_loader(given_name, dtype, row)
    dataset = reader.state.dataset.data if reader.state.dataset else None

    if row['prepare']:
        new_data['prepare_given'].append(
            PrepareGiven(
                appended=False,
                source='',
                prepare=row['prepare']
            )
        )

    if row['ref']:
        if dtype['type'] in (DataTypeEnum.REF.value, DataTypeEnum.GENERIC.value, DataTypeEnum.BACKREF.value):
            ref_model, ref_props = _parse_property_ref(row['ref'])
            new_data['model'] = get_relative_model_name(dataset, ref_model)

            if dtype['type'] == DataTypeEnum.BACKREF.value:
                if len(ref_props) > 1:
                    raise InvalidBackRefReferenceAmount(backref=reader.name)
                elif len(ref_props) == 1:
                    new_data['refprop'] = ref_props[0]

            else:
                new_data['refprops'] = ref_props

        else:
            # TODO: Detect if ref is a unit or an enum.
            new_data['enum'] = row['ref']

    if dataset or row['source']:
        new_data['external'] = {
            'name': row['source'],
        }

    return new_data


def _string_datatype_handler(reader: PropertyReader, row: dict):
    given_name = row['property']
    reader.name = _clean_up_prop_name(row['property'].split('.')[-1])

    if reader.state.model is None:
        context = reader.state.stack[-1]
        reader.error(
            f"Property {reader.name!r} must be defined in a model context. "
            f"Now it is defined in {context.name!r} {context.type} context."
        )
    existing_data = _check_if_property_already_set(reader, row, given_name)
    if row['type'] == DataTypeEnum.TEXT.value and existing_data:
        reader.error(
            f"Property {reader.name!r} with the same name is already "
            f"defined for this {reader.state.model.name!r} model."
        )

    dtype = _get_type_repr(row['type'])
    dtype = _parse_dtype_string(dtype)
    if dtype['error']:
        reader.error(
            dtype['error']
        )

    new_data = _initial_normal_property_schema(given_name, dtype, row)
    dataset = reader.state.dataset.data if reader.state.dataset else None

    if row['prepare']:
        new_data['prepare_given'].append(
            PrepareGiven(
                appended=False,
                source='',
                prepare=row['prepare']
            )
        )
    if row['ref']:
        new_data['enum'] = row['ref']
    if dataset or row['source']:
        new_data['external'] = {
            'name': row['source'],
        }

    return new_data


def _text_datatype_handler(reader: PropertyReader, row: dict):
    given_name = row['property']
    reader.name = _clean_up_prop_name(row['property'].split('.')[-1])

    if reader.state.model is None:
        context = reader.state.stack[-1]
        reader.error(
            f"Property {reader.name!r} must be defined in a model context. "
            f"Now it is defined in {context.name!r} {context.type} context."
        )
    result = _check_if_property_already_set(reader, row, given_name)
    if not (result and result['explicitly_given'] is False and result['type'] == DataTypeEnum.TEXT.value or not result):
        reader.error(
            f"Property {reader.name!r} with the same name is already "
            f"defined for this {reader.state.model.name!r} model."
        )
    dtype = _get_type_repr(row['type'])
    dtype = _parse_dtype_string(dtype)
    if dtype['error']:
        reader.error(
            dtype['error']
        )

    new_data = _initial_text_property_schema(given_name, dtype, row)
    dataset = reader.state.dataset.data if reader.state.dataset else None

    if row['prepare']:
        new_data['prepare_given'].append(
            PrepareGiven(
                appended=False,
                source='',
                prepare=row['prepare']
            )
        )
    if row['ref']:
        new_data['enum'] = row['ref']
    if dataset or row['source']:
        new_data['external'] = {
            'name': row['source'],
        }
    c_lang_name = f'{row["property"]}@C'
    c_lang_data = _empty_property(_initial_normal_property_schema(c_lang_name, dtype, {
        'property': c_lang_name,
        'access': row['access'],
    }))
    c_lang_data['type'] = DataTypeEnum.STRING.value
    c_lang_data['external'] = new_data['external'] if 'external' in new_data else {}
    if result:
        new_data['langs'] = result['langs']
        if new_data['level'] and int(new_data['level']) <= 3:
            new_data['langs']['C'] = c_lang_data
            if 'external' in new_data and new_data['external']:
                new_data['external'] = {}
        result.update(new_data)
        return result

    if new_data['level'] and int(new_data['level']) <= 3:
        new_data['langs'] = {
            'C': c_lang_data
        }
        if 'external' in new_data and new_data['external']:
            new_data['external'] = {}
    return new_data


def _default_datatype_handler(reader: PropertyReader, row: dict):
    return _datatype_handler(reader, row, _initial_normal_property_schema)


def _array_datatype_handler(reader: PropertyReader, row: dict):
    return _datatype_handler(reader, row, _initial_array_property_schema)


def _partial_datatype_handler(reader: PropertyReader, row: dict):
    return _datatype_handler(reader, row, _initial_partial_property_schema)


def _handle_datatype(reader: PropertyReader, row: dict):
    if row['type'] in DATATYPE_HANDLERS:
        handler = DATATYPE_HANDLERS[row['type']]
    else:
        handler = DATATYPE_HANDLERS['_default']
    return handler(reader, row)


DATATYPE_HANDLERS = {
    "_default": _default_datatype_handler,
    DataTypeEnum._PARTIAL_ARRAY.value: _array_datatype_handler,
    DataTypeEnum.ARRAY.value: _array_datatype_handler,
    DataTypeEnum._ARRAY_BACKREF.value: _array_datatype_handler,
    DataTypeEnum._OBJECT.value: _partial_datatype_handler,
    DataTypeEnum._PARTIAL.value: _partial_datatype_handler,
    DataTypeEnum.REF.value: _partial_datatype_handler,
    DataTypeEnum.BACKREF.value: _partial_datatype_handler,
    DataTypeEnum.TEXT.value: _text_datatype_handler,
    DataTypeEnum.STRING.value: _string_datatype_handler
}


def _get_root_prop(reader: PropertyReader, name: str):
    if name in reader.state.model.data['properties']:
        return reader.state.model.data['properties'][name]
    return None


def _clean_up_prop_name(name: str):
    return name.replace('[]', '').split('@', 1)[0]


def _name_complex(name: str) -> bool:
    return '.' in name or '@' in name or '[]' in name


def _restore_previously_nested_data(prop: dict, existing_prop: dict) -> dict:
    if 'properties' in existing_prop:
        prop['properties'] = existing_prop['properties']
    elif 'items' in existing_prop:
        prop['items'] = existing_prop['items']
    elif 'langs' in existing_prop:
        prop['langs'] = existing_prop['langs']
    return prop


def _combine_parent_with_prop(prop_name: str, prop: dict, parent_prop: dict, full_prop: dict):
    return_name = _clean_up_prop_name(prop['given_name'])
    if parent_prop:
        if parent_prop['type'] in ALLOWED_PARTIAL_TYPES:
            if prop_name in parent_prop['properties']:
                prop = _restore_previously_nested_data(prop, parent_prop['properties'][prop_name])
            parent_prop['properties'][prop_name] = prop
            return_name = _clean_up_prop_name(full_prop['given_name'].split('.')[0])
        elif parent_prop['type'] in ALLOWED_ARRAY_TYPES:
            if parent_prop['items']:
                prop = _restore_previously_nested_data(prop, parent_prop['items'])
            parent_prop['items'] = prop
            return_name = _clean_up_prop_name(full_prop['given_name'].split('.')[0])
        elif parent_prop['type'] == DataTypeEnum.TEXT.value:
            given_name = prop['given_name']
            lang_name = given_name.split('@')[-1] if '@' in given_name else 'C'
            parent_prop['langs'][lang_name] = prop
            return_name = _clean_up_prop_name(full_prop['given_name'].split('.')[0])
        else:
            full_prop.clear()
            full_prop.update(prop)
    else:
        if full_prop:
            if prop['type'] in ALLOWED_ARRAY_TYPES:
                prop['items'] = full_prop['items']
            elif prop['type'] in ALLOWED_PARTIAL_TYPES and 'properties' in full_prop:
                prop['properties'] = full_prop['properties']
        full_prop.update(prop)
    return return_name


def _empty_property(data: dict):
    data['explicitly_given'] = False
    return data


def _get_parent_data_array(reader: PropertyReader, given_row: dict, full_name: str, current_parent: dict):
    name = full_name.split('.')[-1]
    array_depth = name.count('[]')
    root_name = name.replace('[]', '')

    empty_array_row = torow(DATASET, {
        'property': full_name,
        'type': DataTypeEnum._PARTIAL_ARRAY.value,
        'access': given_row['access'],
    })

    if not current_parent:
        current_parent.update(_empty_property(_array_datatype_handler(reader, empty_array_row)))

    if given_row.get('type') == DataTypeEnum.BACKREF.value:
        current_parent['type'] = DataTypeEnum._ARRAY_BACKREF.value

    adjustment = 1 if current_parent.get('type') in ALLOWED_ARRAY_TYPES else 0

    for _ in range(array_depth - adjustment):
        current_type = current_parent.get('type')

        if current_type in ALLOWED_ARRAY_TYPES:
            current_parent = _process_allowed_array_type(reader, current_parent, empty_array_row)
        elif current_type in ALLOWED_PARTIAL_TYPES:
            current_parent = _process_allowed_partial_type(reader, current_parent, root_name, empty_array_row)
        else:
            raise NestedDataTypeMismatch(initial=current_type, required=DataTypeEnum.ARRAY.value)

    return current_parent


def _process_allowed_array_type(reader: PropertyReader, current_parent: dict, empty_array_row: dict) -> dict:
    if current_parent.get('items') and current_parent['items'].get('type') not in ALLOWED_ARRAY_TYPES:
        raise NestedDataTypeMismatch(initial=current_parent['type'], required=DataTypeEnum.ARRAY.value)
    if not current_parent.get('items'):
        current_parent['items'] = _empty_property(_array_datatype_handler(reader, empty_array_row))

    return current_parent['items']


def _process_allowed_partial_type(reader: PropertyReader, current_parent: dict, root_name: str,
                                  empty_array_row: dict) -> dict:
    properties = current_parent.setdefault('properties', {})

    if root_name in properties:
        prop_type = properties[root_name].get('type')
        if prop_type not in ALLOWED_ARRAY_TYPES:
            raise NestedDataTypeMismatch(initial=current_parent['type'], required=DataTypeEnum.ARRAY.value)
    else:
        properties[root_name] = _empty_property(_array_datatype_handler(reader, empty_array_row))

    return properties[root_name]


def _get_parent_data_partial(reader: PropertyReader, given_row: dict, full_name: str, current_parent: dict):
    empty_partial_row = torow(DATASET, {
        'property': full_name,
        'type': 'partial',
        'access': given_row['access'],
    })
    name = _clean_up_prop_name(full_name.split('.')[-1])
    if not current_parent:
        current_parent.update(_empty_property(_partial_datatype_handler(reader, empty_partial_row)))
        return current_parent

    if current_parent['type'] in ALLOWED_ARRAY_TYPES:
        if current_parent['items'] and current_parent['items']['type'] not in ALLOWED_PARTIAL_TYPES:
            raise NestedDataTypeMismatch(initial=current_parent['type'], required=DataTypeEnum._PARTIAL.value)
        elif not current_parent['items']:
            current_parent['items'].update(_empty_property(_partial_datatype_handler(reader, empty_partial_row)))
        current_parent = current_parent['items']
    elif current_parent['type'] in ALLOWED_PARTIAL_TYPES:
        if (
            name in current_parent['properties']
            and current_parent['properties'][name]['type'] not in ALLOWED_PARTIAL_TYPES
        ):
            raise NestedDataTypeMismatch(initial=current_parent['type'], required=DataTypeEnum._PARTIAL.value)
        elif name not in current_parent['properties']:
            current_parent['properties'][name] = _empty_property(
                _partial_datatype_handler(reader, empty_partial_row))
        current_parent = current_parent['properties'][name]
    return current_parent


def _get_parent_data_text(reader: PropertyReader, given_row: dict, full_name: str, current_parent: dict):
    split = full_name.split('.')
    name = _clean_up_prop_name(split[-1])
    empty_text_row = torow(DATASET, {
        'property': name if len(split) == 1 else '.'.join(split[:-1] + [name]),
        'type': 'text',
        'access': given_row['access']
    })

    if not current_parent:
        if given_row['type'] == 'text':
            return current_parent

        current_parent.update(_empty_property(_text_datatype_handler(reader, empty_text_row)))
        return current_parent

    if current_parent['type'] in ALLOWED_ARRAY_TYPES:
        if current_parent['items'] and current_parent['items']['type'] != DataTypeEnum.TEXT.value:
            raise NestedDataTypeMismatch(initial=current_parent['properties'][name]['type'], required=DataTypeEnum.TEXT.value)
        elif not current_parent['items']:
            current_parent['items'].update(_empty_property(_text_datatype_handler(reader, empty_text_row)))
        current_parent = current_parent['items']
    elif current_parent['type'] in ALLOWED_PARTIAL_TYPES:
        if name in current_parent['properties'] and current_parent['properties'][name]['type'] != DataTypeEnum.TEXT.value:
            raise NestedDataTypeMismatch(initial=current_parent['properties'][name]['type'], required=DataTypeEnum.TEXT.value)
        elif name not in current_parent['properties']:
            current_parent['properties'][name] = _empty_property(
                _text_datatype_handler(reader, empty_text_row))
        current_parent = current_parent['properties'][name]
    return current_parent


def _extract_and_create_parent_data(
    reader: PropertyReader,
    current_row: dict,
    property_path: str
) -> (dict, dict, str):
    property_parts = property_path.split('.')
    complete_structure = {}
    root_property = _get_root_prop(reader, _clean_up_prop_name(property_parts[0]))

    if root_property:
        complete_structure.update(root_property)

    current_nested_dict = None
    total_parts = len(property_parts)
    property_name = property_path
    accumulated_path = []

    for index, part in enumerate(property_parts):
        accumulated_path.append(part)

        if '[]' in part:
            current_nested_dict = _get_parent_data_array(
                reader, current_row, '.'.join(accumulated_path), current_nested_dict or complete_structure
            )
        elif '@' in part:
            current_nested_dict = _get_parent_data_text(
                reader, current_row, '.'.join(accumulated_path), current_nested_dict or complete_structure
            )

        if index < total_parts - 1:
            if not current_nested_dict and complete_structure:
                current_nested_dict = complete_structure
            else:
                current_nested_dict = _get_parent_data_partial(
                    reader, current_row, '.'.join(accumulated_path), current_nested_dict or complete_structure
                )
        else:
            property_name = _clean_up_prop_name(part)

        if current_nested_dict and current_nested_dict['type'] not in ALLOWED_NESTING_TYPES:
            raise DataTypeCannotBeUsedForNesting(dtype=current_nested_dict['type'])

    return complete_structure, current_nested_dict, property_name


def _get_parent_data(reader: PropertyReader, given_row: dict, name: str):
    split_props = name.split('.')
    full_prop = {}
    root = _get_root_prop(reader, _clean_up_prop_name(split_props[0]))
    if root:
        full_prop.update(root)

    current_parent = None
    count = len(split_props)
    prop_name = name
    full_name = []
    for i, prop in enumerate(split_props):
        full_name.append(prop)
        if '[]' in prop:
            current_parent = _get_parent_data_array(reader, given_row, '.'.join(full_name), current_parent or full_prop)
        if i + 1 != count:
            if not current_parent and full_prop:
                current_parent = full_prop
            else:
                current_parent = _get_parent_data_partial(reader, given_row, '.'.join(full_name),
                                                          current_parent or full_prop)
        else:
            prop_name = _clean_up_prop_name(prop)
        if current_parent and (
            current_parent['type'] not in ALLOWED_PARTIAL_TYPES and current_parent['type'] not in ALLOWED_ARRAY_TYPES):
            raise DataTypeCannotBeUsedForNesting(dtype=current_parent['type'])
    return full_prop, current_parent, prop_name


def _extract_children_from_nested(base: dict, children_name: str) -> dict:
    if base['type'] in ALLOWED_ARRAY_TYPES:
        base = base['items']
    elif base['type'] in ALLOWED_PARTIAL_TYPES:
        base = base['properties'].get(children_name, None)
    elif base['type'] == DataTypeEnum.TEXT.value:
        base = base['langs'].get(children_name, None)
    else:
        raise DataTypeCannotBeUsedForNesting(dtype=base['type'])

    return base


def _check_if_property_already_set(reader: PropertyReader, given_row: dict, full_name: str):
    # Treat '@' as normal '.', since '_extract_children_from_nested' is able to extract based on type
    split = full_name.replace('@', '.').split('.')
    base = {}

    properties = reader.state.model.data['properties']
    root = True
    for name in split:

        base_name = name

        if not base and root:
            skip = True
            root = False
            if _name_complex(name):
                skip = False
                base_name = _clean_up_prop_name(name)

            if base_name not in properties:
                return

            base = properties[base_name]

            if skip:
                continue

        if not base:
            return

        if base.get('given_name', None) == full_name:
            break

        if '[]' in name:
            count = name.count('[]') + 1
            name = name.replace('[]', '')

            for _ in range(count):
                if not base:
                    return
                base = _extract_children_from_nested(base, name)
        else:
            base = _extract_children_from_nested(base, name)
    if (
        base
        and base['given_name'] == full_name
        and base['explicitly_given']
    ):
        reader.error(
            f"Property {full_name!r} with the same name is already "
            f"defined for this {reader.state.model.name!r} model."
        )

    if (
        base
        and ((base['type'] in ALLOWED_PARTIAL_TYPES and given_row['type'] not in ALLOWED_PARTIAL_TYPES)
             or (base['type'] in ALLOWED_ARRAY_TYPES and given_row['type'] not in ALLOWED_ARRAY_TYPES))
    ):
        raise DataTypeCannotBeUsedForNesting(dtype=given_row['type'])
    return base


class AppendReader(TabularReader):
    type: str = 'append'
    data: ManifestRow

    def read(self, row: ManifestRow) -> None:
        self.name = row[REF]
        self.data = row

    def release(self, reader: TabularReader = None) -> bool:
        return True

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        self.state.stack[-1].append(self.data)


class PrefixReader(TabularReader):
    type: str = 'prefix'
    data: PrefixRow

    def read(self, row: Dict[str, str]) -> None:
        if not row['ref']:
            # `ref` is a required parameter.
            return

        self.name = row['ref']

        node = (
            self.state.prop or
            self.state.model or
            self.state.base or
            self.state.resource or
            self.state.dataset or
            self.state.manifest
        )

        if 'prefixes' not in node.data:
            node.data['prefixes'] = {}

        prefixes = node.data['prefixes']

        if self.name in prefixes:
            self.error(
                f"Prefix {self.name!r} with the same name is already "
                f"defined for this {node.name!r} {node.type}."
            )

        self.data = {
            'id': row['id'],
            'eid': f'{self.path}:{self.line}',
            'type': self.type,
            'name': self.name,
            'uri': row['uri'],
            'title': row['title'],
            'description': row['description'],
        }

        prefixes[self.name] = self.data

    def append(self, row: Dict[str, str]) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, AppendReader)

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass


class NamespaceReader(TabularReader):
    type: str = 'ns'
    appendable: bool = True

    def read(self, row: Dict[str, str]) -> None:
        if not row['ref']:
            # `ref` is a required parameter.
            return

        self.name = row['ref']

        manifest = self.state.manifest

        if self.name in manifest.namespaces:
            self.error(
                f"Namespace {self.name!r} with the same name is already "
                f"defined."
            )

        manifest.namespaces.add(self.name)

        self.rows.append({
            'id': row['id'],
            'type': self.type,
            'name': self.name,
            'title': row['title'],
            'description': row['description'],
        })

    def append(self, row: Dict[str, str]) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, AppendReader)

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass


class ParamReader(TabularReader):
    type: str = 'param'
    data: ParamRow
    name: str = None

    def _get_node(self) -> TabularReader:
        return (
            self.state.prop or
            self.state.model or
            self.state.base or
            self.state.resource or
            self.state.dataset or
            self.state.manifest
        )

    def _get_data(self, name: str, row: ManifestRow):
        return {
            'name': name,
            'source': [row[SOURCE]],
            'prepare': [_parse_spyna(self, row[PREPARE])],
            'title': row[TITLE],
            'description': row[DESCRIPTION],
        }

    def _get_and_append_data(self, old: Dict, row: ManifestRow):
        source = row[SOURCE]
        prepare = row[PREPARE]
        if source or prepare:
            old["source"].append(source)
            old["prepare"].append(_parse_spyna(self, prepare))

    def _ensure_params_list(self, node: TabularReader, name: str) -> None:
        if 'params' not in node.data:
            node.data['params'] = {}

    def _check_param_name(self, node: TabularReader, name: str) -> None:
        if 'params' in node.data and name in node.data['params']:
            self.error(
                f"Parameter {name!r} with the same name already defined!"
            )

    def read(self, row: ManifestRow) -> None:
        node = self._get_node()

        self.name = row[REF]
        if not self.name:
            self.error("Parameter must have a name.")

        self._check_param_name(node, self.name)
        self._ensure_params_list(node, self.name)

        self.data = self._get_data(self.name, row)
        node.data['params'][self.name] = self.data

    def append(self, row: ManifestRow) -> None:
        node = self._get_node()

        if row[REF]:
            self.name = row[REF]
            self._check_param_name(node, self.name)
            self._ensure_params_list(node, self.name)

        self._get_and_append_data(node.data['params'][self.name], row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, (AppendReader, LangReader))

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass


class EnumReader(TabularReader):
    type: str = 'enum'
    data: EnumRow
    name: str = None

    def read(self, row: ManifestRow) -> None:
        if row[REF]:
            self.name = row[REF]
        else:
            self.name = self.name or ''

        if not any([
            row[SOURCE],
            row[PREPARE],
            row[ACCESS],
            row[TITLE],
            row[DESCRIPTION],
        ]):
            return

        # FIXME AST should be handled by Env
        source = str(row[SOURCE])
        if not source:
            prepare = _parse_spyna(self, row[PREPARE])
            if isinstance(prepare, dict):
                if prepare['name'] == 'negative':
                    prepare = -prepare['args'][0]
                else:
                    prepare = row[PREPARE]
            source = str(prepare)

        if not source:
            self.error(
                "At least source or prepare must be specified for an enum."
            )

        self.data = {
            'id': row[ID],
            'name': self.name,
            'source': row[SOURCE],
            'prepare': _parse_spyna(self, row[PREPARE]),
            'access': row[ACCESS],
            'title': row[TITLE],
            'description': row[DESCRIPTION],
            'level': row[LEVEL],
        }

        node_data: PropertyRow = self._get_node_data(row)

        if 'enums' not in node_data:
            node_data['enums'] = {}

        if self.name not in node_data['enums']:
            node_data['enums'][self.name] = {}

        enum = node_data['enums'][self.name]

        if source in enum:
            self.error(
                f"Enum {self.name!r} item {source!r} with the same value is "
                f"already defined."
            )
        enum[source] = self.data

    def append(self, row: ManifestRow) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, (AppendReader, LangReader))

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass

    def _get_node_data(self, row: ManifestRow) -> PropertyRow:
        node: TabularReader = (
            self.state.prop
            or self.state.model
            or self.state.base
            or self.state.resource
            or self.state.dataset
            or self.state.manifest
        )

        node_data: PropertyRow = node.data

        if isinstance(node, PropertyReader):
            if node.path_to_current_prop:
                for key in node.path_to_current_prop.split('.'):
                    node_data = node_data[key]

        return node_data


class LangReader(TabularReader):
    type: str = 'lang'

    def read(self, row: ManifestRow) -> None:
        reader = self.state.stack[-1]
        if not isinstance(reader, (
            DatasetReader,
            ResourceReader,
            BaseReader,
            ModelReader,
            PropertyReader,
            EnumReader,
            LangReader
        )):
            self.error(f'Language metadata is not supported on {reader.type}.')
            return

        if 'lang' not in reader.data:
            reader.data['lang'] = {}

        lang = reader.data['lang']

        self.name = row[REF]

        if self.name in lang:
            self.error(
                f"Language {self.name!r} with the same name is already "
                f"defined for this {reader.name!r} {reader.type}."
            )

        lang[self.name] = {
            'id': row[ID],
            'eid': f'{self.path}:{self.line}',
            'type': self.type,
            'ref': self.name,
            'title': row[TITLE],
            'description': row[DESCRIPTION],
        }

    def append(self, row: ManifestRow) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, AppendReader)

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass


class UniqueReader(TabularReader):
    type: str = 'unique'

    def read(self, row: ManifestRow) -> None:
        self.name = row[REF]
        reader = self.state.stack[-1]

        if not isinstance(reader, (
            ModelReader,
            UniqueReader,
            AppendReader
        )):
            self.error(f'Unique reader is not supported for {reader.type}.')
            return

        if self.type not in reader.data:
            reader.data['unique'] = []

    def append(self, row: ManifestRow) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, AppendReader)

    def enter(self) -> None:
        data = [row.strip() for row in self.name.split(',')]
        if data not in self.state.model.data['unique']:
            self.state.model.data['unique'].append(data)

    def leave(self) -> None:
        pass


class CommentReader(TabularReader):
    type: str = 'comment'
    data: CommentData

    def read(self, row: ManifestRow) -> None:
        reader = self.state.stack[-1]

        if 'comments' not in reader.data:
            reader.data['comments'] = []

        comments = reader.data['comments']

        comments.append({
            'id': row[ID],
            'parent': row[REF],
            'author': row[SOURCE],
            'access': row[ACCESS],
            # TODO: parse datetime
            'created': row[TITLE],
            'comment': row[DESCRIPTION],
        })

    def append(self, row: ManifestRow) -> None:
        self.read(row)

    def release(self, reader: TabularReader = None) -> bool:
        return not isinstance(reader, AppendReader)

    def enter(self) -> None:
        pass

    def leave(self) -> None:
        pass


READERS = {
    # Main dimensions
    'dataset': DatasetReader,
    'resource': ResourceReader,
    'base': BaseReader,
    'model': ModelReader,
    'property': PropertyReader,

    # Extra dimensions
    '': AppendReader,
    'prefix': PrefixReader,
    'ns': NamespaceReader,
    'param': ParamReader,
    'enum': EnumReader,
    'lang': LangReader,
    'comment': CommentReader,
    'unique': UniqueReader
}


class State:
    stack: List[TabularReader]

    backends: Dict[str, Dict[str, str]] = None

    models: Set[str]

    manifest: ManifestReader = None
    dataset: DatasetReader = None
    resource: ResourceReader = None
    base: BaseReader = None
    model: ModelReader = None
    prop: PropertyReader = None

    rename_duplicates: bool = False

    def __init__(self):
        self.stack = []
        self.models = set()

    def release(self, reader: TabularReader = None) -> Iterator[ParsedRow]:
        for parent in list(reversed(self.stack)):
            if parent.release(reader):
                if isinstance(parent, (
                    ManifestReader,
                    NamespaceReader,
                    DatasetReader,
                    ModelReader,
                )) and parent.name != "/":
                    yield from parent.items()
                self.stack.pop()
                parent.leave()
            else:
                break

        if reader:
            reader.enter()
            self.stack.append(reader)


def _read_tabular_manifest_rows(
    path: Optional[str],
    rows: Iterator[Tuple[str, List[str]]],
    *,
    rename_duplicates: bool = True,
    allow_updates: bool = False
) -> Iterator[ParsedRow]:
    _, header = next(rows, (None, None))
    if header is None:
        # Looks like an empty file.
        return
    header = _detect_header(path, 1, header)

    defaults = {k: '' for k in MANIFEST_COLUMNS}

    state = State()
    state.rename_duplicates = rename_duplicates
    reader = ManifestReader(state, path, '1', allow_updates=allow_updates)
    reader.read({})
    yield from state.release(reader)

    for line, row in rows:
        _check_row_size(path, line, header, row)
        row = dict(zip(header, row))
        row = {**defaults, **row}
        dimension = _detect_dimension(path, line, row)
        Reader = READERS[dimension]
        reader = Reader(state, path, line, allow_updates=allow_updates)
        reader.read(row)
        yield from state.release(reader)

    yield from state.release()


def _check_row_size(
    path: Optional[str],
    line: int,
    header: List[str],
    row: List[str],
):
    if len(header) != len(row):
        header = ['header', 'row']
        table = [
            ['∅' if x is None else x for x in v]
            for v in zip_longest(header, row)
        ]
        table = tabulate(table, headers=['header', 'row'])
        raise TabularManifestError(
            f"{path}:{line}: "
            "Number of row cells do not match table header, see what is "
            "missing, missing cells marked with ∅ symbol:\n"
            f"{table}"
        )


def read_tabular_manifest(
    format_: TabularFormat = None,
    *,
    path: str = None,
    file: IO = None,
    rename_duplicates: bool = False,
) -> Iterator[ParsedRow]:
    if format_ == TabularFormat.GSHEETS:
        rows = read_gsheets_manifest(path)
    elif format_ == TabularFormat.CSV:
        rows = _read_csv_manifest(path, file)
    elif format_ == TabularFormat.ASCII:
        rows = _read_txt_manifest(path, file)
    elif format_ == TabularFormat.XLSX:
        rows = _read_xlsx_manifest(path)
    else:
        raise ValueError(f"Unknown tabular manifest format {format_!r}.")

    try:
        yield from _read_tabular_manifest_rows(
            path,
            rows,
            rename_duplicates=rename_duplicates,
        )
    except Exception:
        if isinstance(rows, types.GeneratorType):
            rows.close()
        raise


def _read_txt_manifest(
    path: str,
    file: IO[str] = None,
) -> Iterator[Tuple[str, List[str]]]:
    if file:
        yield from _read_ascii_tabular_manifest(file)
    else:
        with pathlib.Path(path).open(encoding='utf-8-sig') as f:
            yield from _read_ascii_tabular_manifest(f)


def _read_csv_manifest(
    path: str,
    file: IO[str] = None,
) -> Iterator[Tuple[str, List[str]]]:
    if file:
        rows = csv.reader(file)
        for i, row in enumerate(rows, 1):
            yield str(i), row
    else:
        with pathlib.Path(path).open(encoding='utf-8') as f:
            rows = csv.reader(f)
            for i, row in enumerate(rows, 1):
                yield str(i), row


def _empty_rows_counter():
    empty_rows = 0

    def _counter(row):
        nonlocal empty_rows
        if any(row):
            empty_rows = 0
        else:
            empty_rows += 1
        return empty_rows

    return _counter


def _read_xlsx_manifest(path: str) -> Iterator[Tuple[str, List[str]]]:
    wb = openpyxl.load_workbook(path)

    yield '1', DATASET

    for sheet in wb:
        rows = sheet.iter_rows(values_only=True)
        cols = next(rows, None)
        if cols is None:
            continue
        cols = normalizes_columns(cols)
        cols = [cols.index(c) if c in cols else None for c in DATASET]

        empty_rows = _empty_rows_counter()
        for i, row in enumerate(rows, 2):
            row = [row[c] if c is not None else None for c in cols]
            yield f'{sheet.title}:{i}', row

            if empty_rows(row) > 100:
                log.warning(
                    f"Too many consequent empty rows, stop reading {path} "
                    f"at {i} row."
                )
                break


def striptable(table):
    return textwrap.dedent(table).strip()


def _join_escapes(row: List[str]) -> List[str]:
    res = []
    for v in row:
        if res and res[-1] and res[-1].endswith('\\'):
            res[-1] = res[-1][:-1] + '|' + v
        else:
            res.append(v)
    return res


def _read_ascii_tabular_manifest(
    lines: Iterable[str],
    *,
    check_column_names: bool = True,
) -> Iterator[Tuple[str, List[str]]]:
    lines = (line.strip() for line in lines)
    lines = filter(None, lines)

    # Read header
    header = next(lines, None)
    if header is None:
        return
    header = normalizes_columns(
        header.split('|'),
        check_column_names=check_column_names,
    )
    yield '1', header

    # Find index where dimension columns end.
    dim = sum(1 for h in header if h in DATASET[:6])
    for i, line in enumerate(lines, 2):
        row = _join_escapes(line.split('|'))
        row = [x.strip() for x in row]
        row = row[:len(header)]
        rem = len(header) - len(row)
        row = row[:dim - rem] + [''] * rem + row[dim - rem:]
        assert len(header) == len(row), line
        yield str(i), row


def read_ascii_tabular_rows(
    manifest: str,
    *,
    strip: bool = False,
    check_column_names: bool = True,
) -> Iterator[List[str]]:
    if strip:
        manifest = striptable(manifest)
    rows = _read_ascii_tabular_manifest(
        manifest.splitlines(),
        check_column_names=check_column_names,
    )
    for line, row in rows:
        yield row


def read_ascii_tabular_manifest(
    manifest: str,
    *,
    strip: bool = False,
    rename_duplicates: bool = False,
) -> Iterator[ParsedRow]:
    if strip:
        manifest = striptable(manifest)
    rows = _read_ascii_tabular_manifest(manifest.splitlines())
    yield from _read_tabular_manifest_rows(
        None,
        rows,
        rename_duplicates=rename_duplicates,
    )


def load_ascii_tabular_manifest(
    context: Context,
    manifest: Manifest,
    manifest_ascii_table: str,
    *,
    strip: bool = False,
) -> None:
    schemas = read_ascii_tabular_manifest(manifest_ascii_table, strip=strip)
    load_manifest_nodes(context, manifest, schemas)
    commands.link(context, manifest)


def get_relative_model_name(dataset: [str, dict], name: str) -> str:
    if isinstance(dataset, str):
        return name.replace(dataset, '')
    if name.startswith('/'):
        return name[1:]
    elif '/' in name:
        return name
    elif dataset is None:
        return name
    else:
        return '/'.join([
            dataset['name'],
            name,
        ])


def to_relative_model_name(model: Model, dataset: Dataset = None) -> str:
    """Convert absolute model `name` to relative."""
    if dataset is None:
        return model.name
    if model.name == f'{dataset.name}/{model.basename}':
        return model.basename
    else:
        return '/' + model.name


def tabular_eid(model: Model):
    if isinstance(model.eid, int):
        return model.eid
    else:
        return 0


class OrderBy(NamedTuple):
    func: Callable[[Union[Dataset, Model, Property, EnumItem]], Any]
    reverse: bool = False


def _order_datasets_by_access(dataset: Dataset):
    return dataset.access or Access.private


def _order_datasets_by_name(dataset: Dataset):
    return dataset.name


DATASETS_ORDER_BY = {
    'access': OrderBy(_order_datasets_by_access, reverse=True),
    'default': OrderBy(_order_datasets_by_name),
}


def _order_models_by_access(model: Model):
    return model.access or Access.private


MODELS_ORDER_BY = {
    'access': OrderBy(_order_models_by_access, reverse=True),
    'default': OrderBy(tabular_eid),
}


def _order_properties_by_access(prop: Property):
    return prop.access or Access.private


PROPERTIES_ORDER_BY = {
    'access': OrderBy(_order_properties_by_access, reverse=True),
}

T = TypeVar('T', Dataset, Model, Property, EnumItem)


def sort(
    ordering: Dict[str, OrderBy],
    items: Iterable[T],
    order_by: Optional[str],
) -> Iterable[T]:
    order: Optional[OrderBy] = None

    if order_by:
        order = ordering[order_by]
    elif 'default' in ordering:
        order = ordering['default']

    if order:
        return sorted(items, key=order.func, reverse=order.reverse)
    else:
        return items


def _end_marker(name):
    yield torow(DATASET, {
        name: "/"
    })


def _prefixes_to_tabular(
    prefixes: Dict[str, UriPrefix],
    *,
    separator: bool = False,
) -> Iterator[ManifestRow]:
    first = True
    for name, prefix in prefixes.items():
        yield torow(DATASET, {
            'id': prefix.id,
            'type': prefix.type if first else '',
            'ref': name,
            'uri': prefix.uri,
            'title': prefix.title,
            'description': prefix.description,
        })
        first = False

    if separator and prefixes:
        yield torow(DATASET, {})


def _backends_to_tabular(
    backends: Dict[str, Backend],
    *,
    separator: bool = False,
) -> Iterator[ManifestRow]:
    for name, backend in backends.items():
        yield torow(DATASET, {
            'id': backend.config.get('id'),
            'type': backend.type,
            'resource': name,
            'source': backend.config.get('dsn'),
        })

    if separator and backends:
        yield torow(DATASET, {})


def _namespaces_to_tabular(
    namespaces: Dict[str, Namespace],
    *,
    separator: bool = False,
) -> Iterator[ManifestRow]:
    namespaces = {
        k: ns
        for k, ns in namespaces.items() if not ns.generated
    }
    first = True
    for name, ns in namespaces.items():
        yield torow(DATASET, {
            'id': ns.id,
            'type': ns.type if first else '',
            'ref': name,
            'title': ns.title,
            'description': ns.description,
        })
        first = False

    if separator and namespaces:
        yield torow(DATASET, {})


def _order_enums_by_access(item: EnumItem):
    return item.access or Access.private


ENUMS_ORDER_BY = {
    'access': OrderBy(_order_enums_by_access, reverse=True),
}


def _enums_to_tabular(
    enums: Optional[Enums],
    *,
    external: bool = True,
    access: Access = Access.private,
    order_by: ManifestColumn = None,
    separator: bool = False,
) -> Iterator[ManifestRow]:
    if enums is None:
        return
    for name, enum in enums.items():
        first = True
        items = sort(ENUMS_ORDER_BY, enum.values(), order_by)
        for item in items:
            if item.access is not None and item.access < access:
                continue
            yield torow(DATASET, {
                'id': item.id,
                'type': 'enum' if first else '',
                'ref': name if first else '',
                'source': item.source if external else '',
                'prepare': unparse(item.prepare),
                'access': item.given.access,
                'title': item.title,
                'description': item.description,
                'level': item.level,
            })
            if lang := list(_lang_to_tabular(item.lang)):
                first = True
                yield from lang
            else:
                first = False

    if separator and enums:
        yield torow(DATASET, {})


def _lang_to_tabular(
    lang: Optional[LangData],
) -> Iterator[ManifestRow]:
    if lang is None:
        return
    first = True
    for name, data in sorted(lang.items(), key=itemgetter(0)):
        yield torow(DATASET, {
            'id': data['id'],
            'type': 'lang' if first else '',
            'ref': name if first else '',
            'title': data['title'],
            'description': data['description'],
        })
        first = False


def _text_to_tabular(
    prop
):
    if not isinstance(prop.dtype, Text):
        return
    for lang in prop.dtype.langs:
        yield torow(DATASET, {
            'id': prop.id,
            'property': prop.name + '@' + lang,
            'type': prop.dtype.name,
            'level': prop.level.value if prop.level is not None else '',
            'access': prop.given.access
        })


def _comments_to_tabular(
    comments: Optional[List[Comment]],
    *,
    access: Access = Access.private,
) -> Iterator[ManifestRow]:
    if comments is None:
        return
    first = True
    for comment in comments:
        if comment.access < access:
            return
        yield torow(DATASET, {
            'id': comment.id,
            'type': 'comment' if first else '',
            'ref': comment.parent,
            'source': comment.author,
            'access': comment.given.access,
            'title': comment.created,
            'description': comment.comment,
        })
        first = False


def _unique_to_tabular(model_unique_data, hide_list: List) -> Iterator[ManifestRow]:
    if not model_unique_data:
        return
    for row in model_unique_data:
        if row not in hide_list:
            yield torow(DATASET, {
                'type': 'unique',
                'ref': ', '.join([r.name for r in row])
            })


def _params_to_tabular(params: List[Param]) -> Iterator[ManifestRow]:
    if not params:
        return
    for param in params:
        for i, (source, prepare) in enumerate(zip(param.source, param.prepare)):
            if isinstance(prepare, NotAvailable):
                prepare = ''
            else:
                prepare = spyna.unparse(prepare)
            if i == 0:
                yield torow(DATASET, {
                    'type': 'param',
                    'ref': param.name,
                    'source': source,
                    'prepare': prepare,
                    'title': param.title,
                    'description': param.description
                })
            else:
                yield torow(DATASET, {
                    'source': source,
                    'prepare': prepare
                })


def _dataset_to_tabular(
    dataset: Dataset,
    *,
    external: bool = True,
    access: Access = Access.private,
    order_by: ManifestColumn = None,
) -> Iterator[ManifestRow]:
    yield torow(DATASET, {
        'id': dataset.id,
        'dataset': dataset.name,
        'level': dataset.level,
        'access': dataset.given.access,
        'title': dataset.title,
        'description': dataset.description,
    })
    yield from _lang_to_tabular(dataset.lang)
    yield from _prefixes_to_tabular(dataset.prefixes, separator=True)
    yield from _enums_to_tabular(
        dataset.ns.enums,
        external=external,
        access=access,
        order_by=order_by,
    )


def _resource_to_tabular(
    resource: Resource,
    *,
    external: bool = True,
    access: Access = Access.private,
) -> Iterator[ManifestRow]:
    backend = resource.backend
    yield torow(DATASET, {
        'id': resource.id,
        'resource': resource.name,
        'source': resource.external if external else '',
        'prepare': unparse(resource.prepare or NA) if external else '',
        'type': resource.type,
        'ref': (
            backend.name
            if (
                external and
                backend and
                backend.origin != BackendOrigin.resource
            )
            else ''
        ),
        'level': resource.level,
        'access': resource.given.access,
        'title': resource.title,
        'description': resource.description,
    })
    yield from _params_to_tabular(resource.params)
    yield from _comments_to_tabular(resource.comments, access=access)
    yield from _lang_to_tabular(resource.lang)


def _base_to_tabular(
    base: Base,
) -> Iterator[ManifestRow]:
    data = {
        'base': base.name,
        'level': base.level.value if base.level else "",
        'id': base.id,
    }
    if base.pk:
        data['ref'] = ', '.join([pk.place for pk in base.pk])
    yield torow(DATASET, data)
    yield from _lang_to_tabular(base.lang)


def _property_to_tabular(
    prop: Property,
    *,
    external: bool = True,
    access: Access = Access.private,
    order_by: ManifestColumn = None,
) -> Iterator[ManifestRow]:
    if prop.name.startswith('_'):
        return

    if prop.access is not None and prop.access < access:
        return

    data = {
        'property': prop.given.name or prop.name,
        'id': prop.id,
        'type': _get_type_repr(prop.dtype),
        'level': prop.level.value if prop.level else "",
        'access': prop.given.access,
        'uri': prop.uri,
        'title': prop.title,
        'description': prop.description,
    }

    if external and prop.external:
        if isinstance(prop.external, list):
            # data['source'] = ', '.join(x.name for x in prop.external)
            # data['prepare'] = ', '.join(
            #     unparse(x.prepare or NA)
            #     for x in prop.external if x.prepare
            # )
            raise DeprecationWarning(
                "Source can't be a list, use prepare instead."
            )
        elif prop.external:
            data['source'] = prop.external.name
            data['prepare'] = unparse(prop.external.prepare or NA)

    yield_rows = []

    if isinstance(prop.dtype, (Array, ArrayBackRef)):
        yield_array_row = prop.dtype.items
        yield_rows.append(yield_array_row)

    elif isinstance(prop.dtype, (Ref, BackRef)):
        model = prop.model

        if model.external and model.external.dataset:
            data['ref'] = to_relative_model_name(
                prop.dtype.model,
                model.external.dataset,
            )

            if isinstance(prop.dtype, Ref):
                pkeys = prop.dtype.model.external.pkeys
                rkeys = prop.dtype.refprops

                if rkeys and pkeys != rkeys:
                    rkeys = ', '.join([p.place for p in rkeys])
                    data['ref'] += f'[{rkeys}]'
            else:
                rkey = prop.dtype.refprop.place
                if prop.dtype.explicit:
                    data['ref'] += f'[{rkey}]'

        else:
            data['ref'] = prop.dtype.model.name

        if prop.dtype.properties:
            for obj_prop in prop.dtype.properties.values():
                yield_rows.append(obj_prop)

    elif isinstance(prop.dtype, Object):
        for obj_prop in prop.dtype.properties.values():
            yield_rows.append(obj_prop)
    elif isinstance(prop.dtype, Text):
        for lang_prop in prop.dtype.langs.values():
            yield_rows.append(lang_prop)

    elif prop.enum is not None:
        data['ref'] = prop.given.enum
    elif prop.unit is not None:
        data['ref'] = prop.given.unit
    data, prepare_rows = _prepare_to_tabular(data, prop)
    if prop.given.explicit:
        yield torow(DATASET, data)
    yield from prepare_rows
    yield from _comments_to_tabular(prop.comments, access=access)
    yield from _lang_to_tabular(prop.lang)
    yield from _enums_to_tabular(
        prop.enums,
        external=external,
        access=access,
        order_by=order_by,
    )
    if yield_rows:
        for yield_row in yield_rows:
            if yield_row:
                yield from _property_to_tabular(yield_row, external=external, access=access, order_by=order_by)


def _prepare_to_tabular(data, prop):
    prep_rows = []
    if prop.given.prepare:
        data['prepare'] = ''
        for prep in prop.given.prepare:
            if prep['appended']:
                prep_rows.append(torow(DATASET, {
                    'source': prep['source'],
                    'prepare': prep['prepare']
                }))
            else:
                if prop.external:
                    data['prepare'] = prep['prepare']
    return data, prep_rows


def _model_to_tabular(
    model: Model,
    *,
    external: bool = True,
    access: Access = Access.private,
    order_by: ManifestColumn = None,
) -> Iterator[ManifestRow]:
    data = {
        'id': model.id,
        'model': model.name,
        'level': model.level.value if model.level else "",
        'access': model.given.access,
        'title': model.title,
        'description': model.description,
        'uri': model.uri if model.uri else "",
    }
    if model.external and model.external.dataset:
        data['model'] = to_relative_model_name(
            model,
            model.external.dataset,
        )
    if external and model.external:
        data.update({
            'source': model.external.name,
            'prepare': unparse(model.external.prepare or NA),
        })
        if (
            not model.external.unknown_primary_key and
            all(p.access >= access for p in model.external.pkeys)
        ):
            # Add `ref` only if all properties are available in the
            # resulting manifest.
            data['ref'] = ', '.join([
                p.name for p in model.external.pkeys
            ])

    hide_list = []
    if model.external:
        if not model.external.unknown_primary_key:
            hide_list = [model.external.pkeys]
    yield torow(DATASET, data)
    yield from _params_to_tabular(model.params)
    yield from _comments_to_tabular(model.comments, access=access)
    yield from _lang_to_tabular(model.lang)
    yield from _unique_to_tabular(model.unique, hide_list)
    props = sort(PROPERTIES_ORDER_BY, model.properties.values(), order_by)
    for prop in props:
        yield from _property_to_tabular(
            prop,
            external=external,
            access=access,
            order_by=order_by,
        )


def datasets_to_tabular(
    context: Context,
    manifest: Manifest,
    *,
    external: bool = True,   # clean content of source and prepare
    access: Access = Access.private,
    internal: bool = False,  # internal models with _ prefix like _txn
    order_by: ManifestColumn = None,
) -> Iterator[ManifestRow]:
    yield from _prefixes_to_tabular(manifest.prefixes, separator=True)
    yield from _backends_to_tabular(manifest.backends, separator=True)
    yield from _namespaces_to_tabular(commands.get_namespaces(context, manifest), separator=True)
    yield from _enums_to_tabular(
        manifest.enums,
        external=external,
        access=access,
        order_by=order_by,
        separator=True,
    )

    seen_datasets = set()
    dataset = None
    resource = None
    base = None
    models = commands.get_models(context, manifest)
    models = models if internal else take(models)
    models = sort(MODELS_ORDER_BY, models.values(), order_by)

    separator = False
    for model in models:
        if model.access < access:
            continue

        if model.external:
            if dataset is None or (model.external.dataset and dataset.name != model.external.dataset.name):
                dataset = model.external.dataset
                if dataset:
                    seen_datasets.add(dataset.name)
                    resource = None
                    separator = True
                    yield from _dataset_to_tabular(
                        dataset,
                        external=external,
                        access=access,
                        order_by=order_by,
                    )
            elif (
                dataset is not None
                and model.external.dataset is None
            ):
                dataset = None
                resource = None
                base = None
                separator = True
                yield from _end_marker('dataset')

            if external and model.external and model.external.resource and (
                resource is None or
                resource.name != model.external.resource.name
            ):
                resource = model.external.resource
                if resource:
                    separator = True
                    yield from _resource_to_tabular(
                        resource,
                        external=external,
                        access=access,
                    )
            elif (
                external
                and model.external
                and model.external.resource is None
                and dataset is not None
                and resource is not None
            ):
                base = None
                yield from _end_marker('resource')

        if separator:
            yield torow(DATASET, {})
        else:
            separator = False

        if model.base and (not base or not is_base_same(model.base, base)):
            base = model.base
            yield from _base_to_tabular(model.base)
        elif base and not model.base:
            base = None
            yield from _end_marker("base")
        yield from _model_to_tabular(
            model,
            external=external,
            access=access,
            order_by=order_by,
        )

    datasets = sort(DATASETS_ORDER_BY, commands.get_datasets(context, manifest).values(), order_by)
    for dataset in datasets:
        if dataset.name in seen_datasets:
            continue
        yield from _dataset_to_tabular(
            dataset,
            external=external,
            access=access,
            order_by=order_by,
        )
        for resource in dataset.resources.values():
            yield from _resource_to_tabular(resource)


def is_base_same(a: Base, b: Base):
    return a.name == b.name and a.level == b.level and a.pk == b.pk


def torow(keys, values) -> ManifestRow:
    return {k: values.get(k) for k in keys}


def render_tabular_manifest(
    context: Context,
    manifest: Manifest,
    cols: List[ManifestColumn] = None,
    *,
    sizes: Dict[ManifestColumn, int] = None,
) -> str:
    rows = datasets_to_tabular(context, manifest)
    return render_tabular_manifest_rows(rows, cols, sizes=sizes)


def render_tabular_manifest_rows(
    rows: Iterable[ManifestRow],
    cols: List[ManifestColumn] = None,
    *,
    sizes: Dict[ManifestColumn, int] = None,
) -> str:
    cols = cols or MANIFEST_COLUMNS
    hs = 1 if ID in cols else 0  # hierarchical cols start
    he = cols.index(PROPERTY)    # hierarchical cols end
    hsize = 1                      # hierarchical column size
    bsize = 3                      # border size
    if sizes is None:
        sizes = dict(
            [(c, len(c)) for c in cols[:hs]] +
            [(c, 1) for c in cols[hs:he]] +
            [(c, len(c)) for c in cols[he:]]
        )
        rows = list(rows)
        for row in rows:
            for i, col in enumerate(cols):
                val = '' if row[col] is None else str(row[col])
                if col == ID:
                    sizes[col] = 2
                elif i < he:
                    size = (hsize + bsize) * (he - hs - i) + sizes[PROPERTY]
                    if size < len(val):
                        sizes[PROPERTY] += len(val) - size
                elif sizes[col] < len(val):
                    sizes[col] = len(val)

    line = []
    for col in cols:
        size = sizes[col]
        line.append(col[:size].ljust(size))
    lines = [line]

    for row in rows:
        if ID in cols:
            value = row[ID]
            if isinstance(value, uuid.UUID):
                value = str(value)
            line = [value[:2] if value else '  ']
        else:
            line = []

        for i, col in enumerate(cols[hs:he + 1]):
            val = row[col] or ''
            if val:
                depth = i
                break
        else:
            val = ''
            depth = 0

        line += [' ' * hsize] * depth
        size = (hsize + bsize) * (he - hs - depth) + sizes[PROPERTY]
        line += [val.ljust(size)]

        for col in cols[he + 1:]:
            val = '' if row[col] is None else str(row[col])
            val = val.replace('|', '\\|')
            size = sizes[col]
            line.append(val.ljust(size))

        lines.append(line)

    lines = [' | '.join(line) for line in lines]
    lines = [l.rstrip() for l in lines]
    return '\n'.join(lines)


SHORT_NAMES = {
    'd': 'dataset',
    'r': 'resource',
    'b': 'base',
    'm': 'model',
    'p': 'property',
    't': 'type',
}


def normalizes_columns(
    cols: List[str],
    *,
    check_column_names: bool = True,
) -> List[ManifestColumn]:
    result: List[ManifestColumn] = []
    unknown: List[str] = []
    invalid: List[str] = []
    for col in cols:
        col = col or ''
        col = col.strip().lower()
        col = SHORT_NAMES.get(col, col)
        col = cast(ManifestColumn, col)
        if col not in MANIFEST_COLUMNS:
            unknown.append(col)
        else:
            if unknown:
                result += unknown
                invalid += unknown
                unknown = []
            result.append(col)
    if check_column_names and invalid:
        if len(invalid) == 1:
            raise PropertyNotFound(property=invalid[0])
        else:
            raise MultipleErrors(
                PropertyNotFound(property=col) for col in invalid
            )
    return result


def write_tabular_manifest(
    context: Context,
    path: str,
    rows: Union[
        Manifest,
        Iterable[ManifestRow],
        None,
    ] = None,
    cols: List[ManifestColumn] = None,
) -> None:
    cols = cols or DATASET

    if rows is None:
        rows = []
    elif isinstance(rows, Manifest):
        rows = datasets_to_tabular(context, rows)

    rows = ({c: row[c] for c in cols} for row in rows)
    if path.endswith('.csv'):
        write_csv(pathlib.Path(path), rows, cols)
    elif path.endswith('.xlsx'):
        write_xlsx(pathlib.Path(path), rows, cols)
    else:
        raise ValueError(f"Unknown tabular manifest format {path!r}.")


def write_csv(
    path: pathlib.Path,
    rows: Iterator[ManifestRow],
    cols: List[ManifestColumn],
) -> None:
    with path.open('w') as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(
    path: Any,
    rows: Iterator[ManifestRow],
    cols: List[ManifestColumn],
) -> None:
    workbook = xlsxwriter.Workbook(path, {
        'in_memory': True,
        'strings_to_formulas': False,
        'strings_to_urls': False,
    })

    bold = workbook.add_format({'bold': True})

    formats = {
        'id': workbook.add_format({
            'align': 'right',
            'valign': 'top',
        }),
        'dataset': workbook.add_format({
            'bold': True,
            'valign': 'top',
            'font_color': '#127622',
        }),
        'resource': workbook.add_format({
            'valign': 'top',
        }),
        'base': workbook.add_format({
            'valign': 'top',
        }),
        'model': workbook.add_format({
            'bold': True,
            'valign': 'top',
            'font_color': '#127622',
        }),
        'property': workbook.add_format({
            'valign': 'top',
            'font_color': '#127622',
        }),
        'type': workbook.add_format({
            'valign': 'top',
        }),
        'ref': workbook.add_format({
            'valign': 'top',
            'font_color': '#127622',
        }),
        'source': workbook.add_format({
            'valign': 'top',
            'font_color': '#c9211e',
        }),
        'prepare': workbook.add_format({
            'valign': 'top',
            'font_color': '#c9211e',
        }),
        'level': workbook.add_format({
            'valign': 'top',
        }),
        'access': workbook.add_format({
            'valign': 'top',
        }),
        'uri': workbook.add_format({
            'valign': 'top',
            'font_color': '#284f80',
        }),
        'title': workbook.add_format({
            'valign': 'top',
            'text_wrap': True,
        }),
        'description': workbook.add_format({
            'valign': 'top',
            'text_wrap': True,
        }),
    }

    sheet = workbook.add_worksheet()
    sheet.freeze_panes(1, 0)  # Freeze the first row.

    sheet.set_column('A:E', 2)   # id, d, r, b, m
    sheet.set_column('F:F', 20)  # property
    sheet.set_column('I:J', 20)  # source, prepare
    sheet.set_column('N:N', 20)  # title
    sheet.set_column('O:O', 30)  # description

    for j, col in enumerate(cols):
        sheet.write(0, j, col, bold)

    for i, row in enumerate(rows, 1):
        for j, col in enumerate(cols):
            val = row[col]
            fmt = formats.get(col)
            sheet.write(i, j, val, fmt)

    workbook.close()


def _get_state_obj(reader: TabularReader) -> Optional[TabularReader]:
    if reader is None or reader.name == "/":
        return None
    else:
        return reader
